import json
import openpyxl
from pathlib import Path
from sigma_parsing.utils import *

def getstr(var):
    if (isinstance(var, str)):
        return var
    return ""

suffix = ".xlsx"
users, filename = get_json_by_pattern("output/*processed*txt")
oname = get_file_name(filename,suffix)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Users"
ws["A1"] = "Name"
ws["B1"] = "Surname"
ws["C1"] = "vk"
ws["D1"] = "School"
ws["E1"] = "Olymp"
ws["F1"] = "Class"
ws["G1"] = "Result"
row = 2
ids = []
for user in users:
    if (not user["processed"]):
        continue 
    for olymp in user["diplomas"]:
        ws["A"+str(row)] = getstr(user["name"])
        ws["B"+str(row)] = getstr(user["surname"])
        ws["C"+str(row)] = "https://vk.com/id" + str(user["official_page"]["id"])
        ws["D"+str(row)] = getstr(user["school"])
        ws["E"+str(row)] = getstr(olymp["olymp"])
        ws["F"+str(row)] = getstr(olymp["class"])
        ws["G"+str(row)] = getstr(olymp["status"])
        row += 1 

wb.save(oname)
