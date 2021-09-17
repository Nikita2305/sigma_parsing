import json
import openpyxl
from pathlib import Path
from sigma_parsing.utils import *

def get_status(our_id, official_id):
    if (our_id == -1):
        return "not selected but search successfull"
    if (our_id == official_id):
        return "selected and matched"
    return "selected but not match"
    
def find_by_id(members, s_id):
    for member in members:
        for account in member["vk_pages"]:
            if (account["id"] == s_id):
                return account
    return {}

suffix = ".xlsx"
users, filename = get_json_by_pattern("output/*processed*txt")
oname = get_file_name(filename,suffix)

ids = []
for user in users:
    if (user["processed"]):
        ids += [(user["official_page"]["id"], user["vk_id"], user)]
    else:
        if (user["vk_id"] in [account["id"] for account in user["vk_pages"]]):
            ids += [(-1, user["vk_id"], user)]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Vertices"
ws["A1"] = "Id"
ws["B1"] = "Label"
ws["C1"] = "Status"
ws["D1"] = "School"
usr = 2
final_ids = []
for pair in ids:
    if (pair[0] == -1): 
        account = find_by_id(users, pair[1])
    else:
        account = find_by_id(users, pair[0])
    id1 = account["id"]
    if (id1 in final_ids):
        continue
    final_ids += [id1]
    ws["A"+str(usr)] = str(id1)
    ws["B"+str(usr)] = account["first_name"] + " " + account["last_name"]
    ws["C"+str(usr)] = get_status(pair[0], pair[1])
    ws["D"+str(usr)] = pair[2]["school"]
    usr += 1

ws = wb.create_sheet("Edges")
ws["A1"] = "Source"
ws["B1"] = "Target"
ws["C1"] = "Type"
edge = 2
for id1 in final_ids:
    account = find_by_id(users, id1)
    for id2 in account["friends"]:
        if (id2 in final_ids):
            ws["A"+str(edge)] = str(id1)
            ws["B"+str(edge)] = str(id2)
            ws["C"+str(edge)] = "Undirected"
        edge += 1

wb.save(oname)
