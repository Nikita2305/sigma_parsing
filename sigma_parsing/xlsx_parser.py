import openpyxl
from pathlib import Path
import json
import re
from sigma_parsing.utils import *

def format_spaces(string):
    if (isinstance(string, str)):
        return " ".join(string.split())
    return string

def log(string):
    print("__LOG__: " + str(string))

def find_cell_with_pattern(sheet, pattern):
    ret = []
    for i in range(1, 20):
        for j in range(1, 50):
            if (isinstance(sheet.cell(i, j).value, str) and
                    re.search(pattern, format_spaces(sheet.cell(i, j).value.lower()))):
                    # pattern == ):
                ret += [(i, j)]
    return ret

def find_format(sheet):
    # Assuming 'columns' is a global/module-level list of dicts like:
    # [{"field_name": "name", "cell_pattern": "Participant Name"}, ...]
    # Assuming 'log' and 'find_cell_with_pattern' functions are defined elsewhere.

    if not columns:
        log("Warning: The 'columns' configuration list is empty. Cannot determine format.")
        return -1, {}

    # Step 1: Find all cells for each column pattern
    found_cells_by_field = {}  # Stores field_name -> list of (row, col_id)
    all_row_indices_from_all_hits = set()

    for column_spec in columns:
        field_name = column_spec["field_name"]
        pattern = column_spec["cell_pattern"]
        
        # find_cell_with_pattern is expected to return a list of (row_idx, col_idx) tuples
        cells_for_current_pattern = find_cell_with_pattern(sheet, pattern)
        log(f'For field "{field_name}" (pattern "{pattern}"), found cells: {cells_for_current_pattern}')
        
        found_cells_by_field[field_name] = cells_for_current_pattern
        for r, _ in cells_for_current_pattern:
            all_row_indices_from_all_hits.add(r)

    # Step 2: Attempt "Happy Path"
    # Try to find a single row where each field has exactly one cell.
    sorted_candidate_rows = sorted(list(all_row_indices_from_all_hits))
    for target_row in sorted_candidate_rows:
        current_col_map = {}
        possible_solution_this_row = True
        
        for column_spec in columns: # Iterate through all defined columns to ensure coverage
            field_name = column_spec["field_name"]
            
            # Get the list of cells found for this specific field's pattern
            # It's possible a pattern for a field yielded no cells.
            cells_for_field = found_cells_by_field.get(field_name, [])
            
            cells_for_field_in_target_row = [(r, c) for r, c in cells_for_field if r == target_row]
            
            if len(cells_for_field_in_target_row) == 1:
                current_col_map[field_name] = cells_for_field_in_target_row[0][1]  # Store col_id
            else:
                # This field has 0 or >1 cells in the target_row. This row is not a "happy path" solution.
                count = len(cells_for_field_in_target_row)
                log(f'Happy path check (row {target_row}): Field "{field_name}" has {count} cells in this row; expected 1.')
                possible_solution_this_row = False
                break  # Move to the next target_row
        
        # If all column_specs were processed successfully for this target_row
        if possible_solution_this_row and len(current_col_map) == len(columns):
            log(f"Happy path: Found format. All fields have exactly one cell in row {target_row}.")
            return target_row, current_col_map

    # Step 3: Attempt "Kinda Happy Path"
    # This path requires:
    # 1. Each field_name (from columns) must have exactly one cell found for its pattern *overall* (across all rows).
    # 2. The rows of these unique cells must be adjacent (or all the same).
    
    single_cell_overall_per_field_map = {}  # field_name -> (row, col_id)
    all_fields_qualify_for_kinda_happy = True

    if not columns: # Should have been caught, but as a safeguard
        all_fields_qualify_for_kinda_happy = False

    if all_fields_qualify_for_kinda_happy: # Only proceed if columns is not empty
        for column_spec in columns:
            field_name = column_spec["field_name"]
            # Get the list of cells found for this field's pattern (populated in Step 1)
            cells_for_field = found_cells_by_field.get(field_name) # Default is None if key missing

            if cells_for_field is not None and len(cells_for_field) == 1:
                single_cell_overall_per_field_map[field_name] = cells_for_field[0]
            else:
                all_fields_qualify_for_kinda_happy = False
                count = len(cells_for_field) if cells_for_field is not None else 0
                status_msg = "pattern found no cells" if cells_for_field is None else f"pattern found {count} cells"
                log(f'Kinda happy path precondition failed: Field "{field_name}" {status_msg} overall; expected exactly 1 cell for this path.')
                break
    
    # Check if all qualifications met and the map is complete
    if all_fields_qualify_for_kinda_happy and len(single_cell_overall_per_field_map) == len(columns) and columns:
        rows_of_the_single_cells = [cell_data[0] for cell_data in single_cell_overall_per_field_map.values()]
        
        if not rows_of_the_single_cells: # Should not happen if map is populated and columns not empty
             log("Kinda happy path: No rows found from single_cell_overall_per_field_map, though map seemed valid.")
        else:
            min_row = min(rows_of_the_single_cells)
            max_row = max(rows_of_the_single_cells)

            if max_row - min_row <= 1:
                # Condition met: all unique cells (one per field) are on the same row or two adjacent rows.
                result_row_for_kinda_happy = max_row  # "assign second row to result" -> max_row
                
                col_map_for_kinda_happy = {
                    field: cell_data[1]  # cell_data is (row, col_id)
                    for field, cell_data in single_cell_overall_per_field_map.items()
                }
                
                log(f"WARNING: Kinda happy path: Format found. All fields have one cell overall. " +
                    f"Cells span rows {min_row} to {max_row}. Assigning result to row {result_row_for_kinda_happy}.")
                return result_row_for_kinda_happy, col_map_for_kinda_happy
            else:
                log(f"Kinda happy path: Cells (one per field) are too far apart (rows {min_row} to {max_row}).")


    # Step 4: Otherwise (failure)
    log("Format not found: Neither happy path nor kinda happy path conditions were met.")
    return -1, {}

def find_number(string):
    start = -1
    end = -2
    for i in range(len(string)):
        if ('0' <= string[i]  and string[i] <= '9'):
            end = i
            if (start == -1):
                start = i
    return string[start : end + 1]

def find_user_id(userlist, user):
    for i in range(len(userlist)):
        if (userlist[i]["name"] == user["name"] and
                userlist[i]["surname"] == user["surname"]):
            return i
    return -1

def find_olymp(olymplist, olymp):
    for x in olymplist:
        if (x["olymp"] == olymp["olymp"]):
            return True
    return False 

suffix = '.xlsxout.txt'
members_oname = get_file_name("output/members", suffix)

with open(xlsxconfig_iname) as f:
    columns = json.load(f)

xlsx_files = [path for path in Path(xlsx_path).rglob('*.xlsx')]
members = []

for xlsx_file in xlsx_files:
    fname = str(xlsx_file)
    subject = fname[fname.rfind('/')+1:fname.rfind('.xlsx')]
    wb = openpyxl.load_workbook(xlsx_file)
    for sheet in wb:
        description = format_spaces(sheet.title)
        class_number = find_number(description)  
        log(subject + ", class:" + class_number)
        zero_row, column_interpret = find_format(sheet)
        if (zero_row == -1):
            log("Формат не найден")
        else:
            while (True):
                zero_row += 1
                user_dict = {}
                OK = True

                for column in columns:
                    col_number = column_interpret[column["field_name"]]
                    value = format_spaces(sheet.cell(zero_row, col_number).value)
                    if (column["field_name"] == "name" or
                                    column["field_name"] == "surname"):
                        if (not isinstance(value, str) or not re.search(column["data_pattern"], value)):
                            OK = False
                    user_dict[column["field_name"]] = str(value)
                     
                if (not OK):
                    break  
          
                event_dict = {"olymp": subject, "class": class_number, "status": user_dict.get("status")}

                user_id = find_user_id(members, user_dict) 
                if (user_id == -1):   
                    user_dict["diplomas"] = [event_dict]
                    members += [user_dict]
                else:
                    if (not find_olymp(members[user_id]["diplomas"], event_dict)):
                        members[user_id]["diplomas"] += [event_dict]
            log("Обработан")
    print()

print("Обработано: " + str(len(members)))
save_as_json(members, members_oname)
