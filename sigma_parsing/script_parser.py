import openpyxl
from pathlib import Path
import json

def format_spaces(string):
    if (isinstance(string, str)):
        return " ".join(string.split())
    return string

def log(string):
    print("__LOG__: " + str(string))

def find_equal(sheet, string):
    ret = []
    for i in range(1, 50):
        for j in range(1, 50):
            # print(sheet.cell(i, j).value)
            if (isinstance(sheet.cell(i, j).value, str) and format_spaces(sheet.cell(i, j).value.lower()) == string):
                ret += [(i, j)]
    return ret

def find_substring(sheet, string):
    ret = []
    for i in range(1, 50):
        for j in range(1, 50):
            # print(sheet.cell(i, j).value)
            if (isinstance(sheet.cell(i, j).value, str) and format_spaces(sheet.cell(i, j).value.lower()).find(string) != -1):
                ret += [(i, j)]
    return ret

def find_format(sheet):
    dict_ret = {"name": 1, "surname": 1, "patronymic": 1, "school": 1, "status": 1}
    surname_list = find_equal(sheet, "фамилия")
    status_list = find_substring(sheet, "диплом")
    if (len(surname_list) != 1 or
            len(status_list) != 1 or
            surname_list[0][0] != status_list[0][0]):
        log(surname_list)
        log(status_list)
        return -1, {}
    surname_row = surname_list[0][0]
    surname_col = surname_list[0][1]
    dict_ret["surname"] = surname_col
    dict_ret["name"] = surname_col + 1
    dict_ret["patronymic"] = surname_col + 2
    dict_ret["school"] = surname_col + 3
    dict_ret["status"] = status_list[0][1]
    return surname_row, dict_ret 

def find_number(string):
    start = -1
    end = -2
    for i in range(len(string)):
        if ('0' <= string[i]  and string[i] <= '9'):
            end = i
            if (start == -1):
                start = i
    return string[start : end + 1]

def find_user_id(userlist, user):
    for i in range(len(userlist)):
        if (userlist[i]["name"] == user["name"] and
                userlist[i]["surname"] == user["surname"] and
                userlist[i]["patronymic"] == user["patronymic"]):
            return i
    return -1

def find_olymp(olymplist, olymp):
    for x in olymplist:
        if (x["olymp"] == olymp["olymp"]):
            return True
    return False



xlsx_files = [path for path in Path('./temp').rglob('itog_*.xlsx')]
final_list = []
try:
    with open('output/members.txt', 'r') as f:
        final_list = json.load(f)
    log("Найден members.txt\n")
except:
    log("Не найден member.txt, ну и ничего страшного\n")

for xlsx_file in xlsx_files:
    fname = str(xlsx_file)
    subject = fname[fname.find('_')+1:fname.rfind('.xlsx')]
    workbook = openpyxl.load_workbook(xlsx_file)
    for sheet in workbook:
        description = format_spaces(sheet.title)
        class_number = find_number(description)  
        log(subject + ", class:" + class_number)
        zero_row, column_interpret = find_format(sheet)
        if (zero_row == -1):
            log("Формат не найден")
        else:
            while (True):
                zero_row += 1
                user_dict = {}
                if (not isinstance(sheet.cell(zero_row, column_interpret["surname"]).value, str) or
                                        len(sheet.cell(zero_row, column_interpret["surname"]).value) == 0):
                        break
                for col_name in column_interpret.keys():
                    col_number = column_interpret[col_name] 
                    user_dict[col_name] = format_spaces(sheet.cell(zero_row, col_number).value)
                event_dict = {"olymp": subject, "class": class_number, "status": user_dict["status"]}
                user_dict.pop("status")

                user_id = find_user_id(final_list, user_dict) 
                if (user_id == -1):   
                    user_dict["diplomas"] = [event_dict]
                    final_list += [user_dict]
                else:
                    if (not find_olymp(final_list[user_id]["diplomas"], event_dict)):
                        final_list[user_id]["diplomas"] += [event_dict] 
            log("Обработан")
        print()
with open('output/members.txt', 'w') as f:
    print(json.dumps(final_list, ensure_ascii=False, indent=4), file=f)
