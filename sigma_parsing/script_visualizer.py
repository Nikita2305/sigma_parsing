import json
import openpyxl
from pathlib import Path

def getstr(var):
    if (isinstance(var, str)):
        return var
    return ""

def is_official(members, user_id):
    for member in members:
        if (member["processed"] and member["official_page"]["id"] == user_id):
            return True
    return False

with open("output/members.txt") as f:
    users = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Edges"
ws["A1"] = "Source"
ws["B1"] = "Target"
ws["C1"] = "Type"
edge = 2
ids = []
for user in users:
    if (not user["processed"]):
        continue
    account = user["official_page"]
    id1 = account["id"]
    if (id1 in ids):
        continue
    ids += [id1]
    for id2 in account["friends"]:
        if (not is_official(users,id2)):
            continue
        ws["A"+str(edge)] = str(id1)
        ws["B"+str(edge)] = str(id2)
        ws["C"+str(edge)] = "Undirected"
        edge += 1

ws = wb.create_sheet("Vertices")
ws["A1"] = "Id"
ws["B1"] = "Label"
ws["C1"] = "Olymps"
ws["D1"] = "School"
usr = 2
ids = []
for user in users:
    if (not user["processed"]):
        continue
    account = user["official_page"]
    id1 = account["id"]
    if (id1 in ids):
        continue
    ids += [id1]
    ws["A"+str(usr)] = str(id1)
    ws["B"+str(usr)] = getstr(user["name"]) + " " + getstr(user["surname"])
    ws["C"+str(usr)] = str([getstr(olymp["olymp"]) + "(" + getstr(olymp["class"]) + " класс)" for olymp in user["diplomas"]])
    ws["D"+str(usr)] = getstr(user["school"])
    usr += 1
wb.save("output/graph.xlsx")
